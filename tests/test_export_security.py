"""Export-path security regressions.

A dataset that arrives from outside controls both its values *and* its column
names. These tests assert that neither can carry an executable payload into a
spreadsheet, and that a compressed file cannot expand into an unbounded frame.
"""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
from loguru import logger

from dqms.config.settings import Settings
from dqms.core.exceptions import FileTooLargeError
from dqms.services.loader import FileLoader
from dqms.utils.io import export_dataframe
from dqms.utils.logging import configure_logging, get_logger

PAYLOAD = "=cmd|' /C calc'!A0"


def test_csv_header_formula_is_neutralised(tmp_path: Path) -> None:
    frame = pd.DataFrame({PAYLOAD: [1, 2], "ok": [3, 4]})
    destination = export_dataframe(frame, tmp_path / "out.csv")
    header = destination.read_text(encoding="utf-8").splitlines()[0]
    assert not header.lstrip('"').startswith("=")
    assert "'=cmd" in header


def test_excel_header_is_not_written_as_a_live_formula(tmp_path: Path) -> None:
    frame = pd.DataFrame({PAYLOAD: [1, 2]})
    destination = export_dataframe(frame, tmp_path / "out.xlsx")
    sheet = openpyxl.load_workbook(destination).active
    assert sheet is not None
    # openpyxl stores a string starting with '=' as data_type 'f' - a formula.
    assert sheet.cell(row=1, column=1).data_type != "f"


def test_excel_cell_value_is_not_written_as_a_live_formula(tmp_path: Path) -> None:
    frame = pd.DataFrame({"note": [PAYLOAD, "safe"]})
    destination = export_dataframe(frame, tmp_path / "vals.xlsx")
    sheet = openpyxl.load_workbook(destination).active
    assert sheet is not None
    assert sheet.cell(row=2, column=1).data_type != "f"


def test_sanitize_false_leaves_the_payload_intact(tmp_path: Path) -> None:
    """The opt-out must genuinely opt out, so the setting is meaningful."""
    frame = pd.DataFrame({PAYLOAD: [1]})
    destination = export_dataframe(frame, tmp_path / "raw.csv", sanitize=False)
    assert destination.read_text(encoding="utf-8").splitlines()[0].startswith("=")


def test_compressed_file_cannot_exceed_the_memory_budget(
    settings: Settings, tmp_path: Path
) -> None:
    """A tiny Parquet file that expands to a large frame must be refused."""
    path = tmp_path / "bomb.parquet"
    pd.DataFrame({"c": ["A" * 2000] * 20_000}).to_parquet(path, compression="zstd")
    assert path.stat().st_size < 1_000_000  # small on disk

    settings.security.max_frame_memory_mb = 1
    with pytest.raises(FileTooLargeError, match="in-memory"):
        FileLoader(settings).load(path)


def test_memory_budget_allows_ordinary_data(settings: Settings, tmp_path: Path) -> None:
    path = tmp_path / "small.csv"
    pd.DataFrame({"a": range(1000)}).to_csv(path, index=False)
    settings.security.max_frame_memory_mb = 64
    assert len(FileLoader(settings).load(path)) == 1000


def test_newline_in_untrusted_text_cannot_forge_a_log_line(settings: Settings) -> None:
    configure_logging(settings, force=True)
    buffer = io.StringIO()
    sink_id = logger.add(buffer, format="{level} | {message}", colorize=False)
    try:
        get_logger("dqms.test").info(
            "Loading {}", "evil.csv\n2099-01-01 | CRITICAL | forged"
        )
    finally:
        logger.remove(sink_id)

    emitted = [line for line in buffer.getvalue().splitlines() if line.strip()]
    assert len(emitted) == 1
    assert "\\n" in emitted[0]
